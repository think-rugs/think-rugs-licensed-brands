"""
extract_cl.py

Builds the Catherine Lansfield (key "Catherine Lansfield") and Catherine
Lansfield Kids (key "CL Kids") blocks of data/brand_data.json from the supplied
Catherine Lansfield product workbook, which carries both collections in one
sheet, and merges them in, leaving the other brands untouched. Re runnable:
replaces both blocks each time.

Usage: python3 scripts/extract_cl.py [path/to/workbook.xlsx]
Default workbook: source/Catherine_Lansfield_Product_Information.xlsx

Normalisation rules (decided against the source workbook, June 2026):
- The brand split comes from the Description prefix ("CL Kids - " vs "CL - ").
  The workbook's own Range column is unreliable: it labels the kids 80 x 150 and
  120 x 170 rows "Catherine Lansfield" and only the 133 x 133 circles "Catherine
  Lansfield Kids".
- The colourway display name is parsed from the Description middle segment with
  the design name and any trailing "Circle" or "Rug" removed, falling back to
  the Colour column.
- Washable? carries mixed booleans (Yes, No, True, False); Yes and True count as
  washable, and washable designs always carry the Washable feature chip.
- Construction strings that bundle the material ("Digitally Printed 100%
  Polyester", any casing) are split into construction and materials to match the
  rest of the suite. Trailing whitespace ("Printed Flat Woven ") is stripped.
- "Anti Slip backing" is normalised to the suite's "Anti Slip" feature chip, and
  "Easy Care" is folded into "Easy Clean" (the workbook uses both for the same
  designs).
- The Shape typo "Retangle" is corrected, and the 133 x 133 rows whose
  description says Circle are given shape Circle. Shape "Runner" sets the runner
  flag. Sizes like "061 x 170" lose their leading zeros.
- Origin "Turkey" is written "Turkiye" to match the suite convention. Six
  machine woven designs (Cameo Floral, Country Floral, Kelso Check, Larsson Geo,
  Pippa Floral, Twilight Animals) have no origin in the workbook; they are left
  blank, reported on each run, and the brochure hides empty spec rows. Confirm
  and fill these when known.
- The kids 133 x 133 circle rows have no ecommerce copy or EUR prices in the
  workbook; copy is taken from sibling sizes of the same colourway and the EUR
  cells stay empty (the brochure shows GBP only for them).
- Two Cameo Floral colourways (Sage / Green and Yellow) have no ecommerce copy
  on any row. Their titles follow the design's own pattern and their
  descriptions are DRAFTED COPY adapted from a sibling colourway. Review before
  licensor approval; replace from the workbook when official copy lands.
- Design intros are not in the workbook; they are generated from workbook facts
  (construction, washability, colourway count) in a fixed pattern and are
  DRAFTED COPY pending official collection copy.
- Each Catherine Lansfield design carries a collection group for the brochure's
  Woven / Washable organisation: Machine Woven constructions are "Woven", the
  printed flat weaves are "Washable". Designs are ordered Woven first, then
  Washable, alphabetical within each group. The kids designs all carry the
  Washable group, so their brochure shows no grouping (it needs two groups).
- Each colourway carries an aliases list (the workbook's Colour column and
  Colour 1 values where they differ from the display name, for example Navy is
  "Blue" and Terra is "Terracotta" in the photo library). The image matcher
  uses these to map photography file names to product codes.
- Em dashes and en dashes anywhere in carried copy are replaced with commas
  (house copy style).
"""
import json, os, re, sys
from collections import OrderedDict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else f'{ROOT}/source/Catherine_Lansfield_Product_Information.xlsx'

KEY_MAIN, KEY_KIDS = 'Catherine Lansfield', 'CL Kids'


def clean(v):
    s = '' if v is None else str(v).strip()
    s = s.replace('\u2014', ', ').replace('\u2013', ', ')
    return re.sub(r'\s+,', ',', re.sub(r'  +', ' ', s))


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h).split('\n')[0].strip(): i for i, h in enumerate(header) if h}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[idx['Product Code']] is None and r[idx['Variant Code']] is None and r[idx['Description']] is None:
            break
        rows.append(r)

    def g(r, name):
        return clean(r[idx[name]])

    def num(r, name):
        v = r[idx[name]]
        return round(float(v), 2) if v not in (None, '') else None

    # ---- per row normalisation ----
    def norm_row(r):
        desc = g(r, 'Description')
        kids = desc.startswith('CL Kids')
        design = g(r, 'Design')

        # colourway display name from the description middle segment
        mid = re.sub(r'^CL( Kids)?\s*-\s*', '', desc)
        mid = re.split(r'\s*-\s*\d', mid)[0].strip()
        colour = mid[len(design):].strip() if mid.lower().startswith(design.lower()) else ''
        colour = re.sub(r'\s+(Circle|Rug)$', '', colour, flags=re.I).strip()
        if not colour:
            colour = g(r, 'Colour')
        colour = re.sub(r'\s*/\s*', ' / ', colour)

        # construction strings that bundle the material
        con = g(r, 'Construction')
        mat1, mat2 = g(r, 'Material 1'), g(r, 'Material 2')
        m = re.match(r'(.*?)\s*(\d+%\s+\w+)$', con)
        if m:
            con = m.group(1).strip()
            split_mat = m.group(2)
            # prefer the more specific form: "100% Polyester" beats a bare "Polyester"
            if not mat1 or mat1.lower() in split_mat.lower():
                mat1 = split_mat
        con = ' '.join(w if w.isupper() else w.capitalize() for w in con.split())
        fix_fibre = lambda s: s.replace('Polyproplyene', 'Polypropylene')
        materials = fix_fibre(mat1 + (f' & {mat2}' if mat2 else ''))

        shape = g(r, 'Shape').replace('Retangle', 'Rectangle')
        if re.search(r'\bCircle\b', desc, re.I):
            shape = 'Circle'

        size = re.sub(r'\b0+(\d)', r'\1', g(r, 'Size (cm)'))

        feats = []
        for f in ('Feature 1', 'Feature 2', 'Feature 3'):
            v = g(r, f).replace('Anti Slip backing', 'Anti Slip').replace('Easy Care', 'Easy Clean')
            if v and v not in feats:
                feats.append(v)
        washable = g(r, 'Washable?') in ('Yes', 'True')
        if washable and 'Washable' not in feats:
            feats.insert(0, 'Washable')

        return {
            'kids': kids, 'design': design, 'code': g(r, 'Product Code'),
            'variant': g(r, 'Variant Code'), 'colour': colour,
            'colour_col': re.sub(r'\s*/\s*', ' / ', g(r, 'Colour')),
            'c1': g(r, 'Colour 1'), 'c2': g(r, 'Colour 2'),
            'title': g(r, 'Ecommerce Title'), 'desc': g(r, 'Ecommerce Description'),
            'construction': con, 'materials': materials,
            'pile': g(r, 'Pile Height (cm)'),
            'origin': g(r, 'Country of Origin').replace('Turkey', 'Turkiye'),
            'features': feats, 'washable': washable,
            'size': size, 'shape': shape, 'runner': shape == 'Runner',
            'rrp': num(r, 'RRP (GBP)'), 'trade': num(r, 'Wholesale Price (GBP)'),
            'rrpEur': num(r, 'RRP (EURO)'), 'tradeEur': num(r, 'Wholesale Price (EUR)'),
        }

    norm = [norm_row(r) for r in rows]

    # ---- group into designs and colourways, preserving workbook order ----
    def build(part, brand_name):
        designs = OrderedDict()
        for n in part:
            d = designs.setdefault(n['design'], OrderedDict())
            d.setdefault(n['code'], []).append(n)

        out, drafted, no_origin = [], [], []
        for dname, cws in designs.items():
            allrows = [n for v in cws.values() for n in v]
            cons = sorted({n['construction'] for n in allrows if n['construction']})
            # the workbook describes the printed flat weaves two ways; collapse to
            # the specific construction, Digitally Printed stays as a feature chip
            if 'Digitally Printed' in cons and 'Printed Flat Woven' in cons:
                cons = [c for c in cons if c != 'Digitally Printed']
            mats = sorted({n['materials'] for n in allrows if n['materials']})
            # drop vague forms contained in a more specific one ("Polyester" vs "100% Polyester")
            mats = [m for m in mats if not any(m != o and m.lower() in o.lower() for o in mats)]
            piles = sorted({n['pile'] for n in allrows if n['pile']}, key=float)
            origins = sorted({n['origin'] for n in allrows if n['origin']})
            feats = []
            for n in allrows:
                for f in n['features']:
                    if f not in feats:
                        feats.append(f)
            if not origins:
                no_origin.append(dname)

            washable = any(n['washable'] for n in allrows)
            colourways = []
            for code, group in cws.items():
                title = next((n['title'] for n in group if n['title']), '')
                desc = next((n['desc'] for n in group if n['desc']), '')
                if not title or not desc:
                    # draft from a sibling colourway of the same design
                    sib = next((s for cc, gg in cws.items() if cc != code
                                for s in gg if s['title'] and s['desc']), None)
                    colour = group[0]['colour']
                    if not title:
                        title = (sib['title'].replace(sib['colour'], colour) if sib and sib['colour'] in sib['title']
                                 else f'{brand_name} {dname} {colour} ' + ('Washable Rug' if washable else 'Rug'))
                    if not desc and sib:
                        desc = sib['desc'].replace(sib['colour'], colour)
                    if not desc:
                        desc = (f'The {brand_name} {dname} {colour} Rug brings the {dname} design home '
                                f'in a {colour.lower()} colourway.')
                    drafted.append(code)
                aliases = []
                for a in (group[0]['colour_col'], group[0]['c1']):
                    if a and a.lower() != group[0]['colour'].lower() and a not in aliases:
                        aliases.append(a)
                colourways.append({
                    'code': code,
                    'colour': group[0]['colour'],
                    'aliases': aliases,
                    'title': title, 'desc': desc,
                    'c1': group[0]['c1'], 'c2': group[0]['c2'],
                    'sizes': [{
                        'size': n['size'], 'shape': n['shape'], 'runner': n['runner'],
                        'variant': n['variant'],
                        'rrp': n['rrp'], 'trade': n['trade'],
                        'rrpEur': n['rrpEur'], 'tradeEur': n['tradeEur'],
                    } for n in group],
                })

            n_cw = len(colourways)
            cw_word = {1: 'one colourway', 2: 'two colourways', 3: 'three colourways',
                       4: 'four colourways', 5: 'five colourways', 6: 'six colourways'}.get(n_cw, f'{n_cw} colourways')
            con_l = (cons[0].lower() if cons else 'made')
            if washable:
                tail = ' with an anti slip backing' if 'Anti Slip' in feats else ''
                intro = f'{dname} is {con_l} and fully machine washable{tail}, in {cw_word}.'
            elif 'Hand Carved' in feats and 'High Density Weave' in feats:
                intro = f'{dname} is {con_l} with a hand carved, high density pile, in {cw_word}.'
            elif 'Hand Carved' in feats:
                intro = f'{dname} is {con_l} with a hand carved pile, in {cw_word}.'
            else:
                intro = f'{dname} is {con_l}, in {cw_word}.'

            trades = [s['trade'] for c in colourways for s in c['sizes'] if s['trade'] is not None]
            rrps = [s['rrp'] for c in colourways for s in c['sizes'] if s['rrp'] is not None]
            out.append({
                'name': dname,
                'group': 'Woven' if any('Machine Woven' in c for c in cons) else 'Washable',
                'intro': intro,
                'construction': ' / '.join(cons),
                'materials': ' / '.join(mats),
                'pile': piles[0] if len(piles) == 1 else f'{piles[0]} to {piles[-1]}',
                'origin': ' and '.join(origins),
                'features': feats,
                'fromRRP': min(rrps), 'fromTrade': min(trades),
                'colourways': colourways,
            })
        out.sort(key=lambda d: ({'Woven': 0, 'Washable': 1}[d['group']], d['name']))
        return out, drafted, no_origin

    main_rows = [n for n in norm if not n['kids']]
    kids_rows = [n for n in norm if n['kids']]
    main_designs, main_drafted, main_no_origin = build(main_rows, 'Catherine Lansfield')
    kids_designs, kids_drafted, kids_no_origin = build(kids_rows, 'Catherine Lansfield')

    path = f'{ROOT}/data/brand_data.json'
    data = json.load(open(path))
    data[KEY_MAIN] = {'brand': 'Catherine Lansfield', 'designs': main_designs}
    data[KEY_KIDS] = {'brand': 'Catherine Lansfield Kids', 'designs': kids_designs}
    json.dump(data, open(path, 'w'), indent=1, ensure_ascii=False)

    def counts(designs):
        cw = sum(len(d['colourways']) for d in designs)
        v = sum(len(c['sizes']) for d in designs for c in d['colourways'])
        return f'{len(designs)} designs, {cw} colourways, {v} variants'

    print(f'{KEY_MAIN}: {counts(main_designs)}')
    print(f'{KEY_KIDS}: {counts(kids_designs)}')
    if main_drafted or kids_drafted:
        print('drafted copy (review before approval):', main_drafted + kids_drafted)
    if main_no_origin or kids_no_origin:
        print('no origin in workbook (left blank, confirm):', main_no_origin + kids_no_origin)


if __name__ == '__main__':
    main()
