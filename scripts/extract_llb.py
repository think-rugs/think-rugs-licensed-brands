"""
extract_llb.py

Builds the House Llewelyn-Bowen (key "LLB") block of data/brand_data.json from
the Think Rugs x LLB product workbook and merges it in, leaving the other brands
untouched. Re runnable: replaces the LLB block each time.

Usage: python3 scripts/extract_llb.py [path/to/workbook.xlsx]
Default workbook: public/downloads/House_Llewelyn_Bowen_Product_Info.xlsx

Normalisation rules (decided against the source workbook, June 2026):
- Design names are stripped of stray whitespace ("Hedgerovia ", "Sarabande ").
- The colourway display name is parsed from the Description column, which carries
  the marketing names (Claret, Ochre), not the Colour column (Red, Yellow). The
  Colour 1 / Colour 2 columns still drive the colour filter.
- Construction strings that bundle the material ("Digitally Printed 100%
  Polyester", "Machine Made 100% Polypropylene") are split into construction and
  materials to match the rest of the suite.
- Imperial is made in two factories: Bronze, Pearl and Silver from Turkiye (0.8 cm
  pile, 160 x 230 and 200 x 290) and the rest from Egypt (1 cm pile, 160 x 235 and
  200 x 285). The design level pile and origin show both; per size data is exact.
- Origin "Turkey" is written "Turkiye" to match the suite convention.
- "Anti Slip backing" is normalised to the suite's "Anti Slip" feature chip.
- The 61 x 230 size is flagged as a runner. Mayfair's 60 x 120 and 80 x 150 are
  small mats, not runners, and are left unflagged.
- The em dash in the Mayfair ecommerce copy is replaced with a comma (house copy
  style: no em dashes or long hyphens).
- The six washable designs and Dolce Vita have no ecommerce title or description
  in the workbook. Titles follow the brand's own pattern, and descriptions and all
  design intros are DRAFTED COPY written for this brochure. Review before licensor
  approval; replace from the workbook when official copy lands.
"""
import json, os, re, sys
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else f'{ROOT}/public/downloads/House_Llewelyn_Bowen_Product_Info.xlsx'

DESIGN_ORDER = ['Imperial', 'Damask', 'Tropicana', 'Cotswold Fantasia', 'Hedgerovia',
                'Sarabande', 'Dandy', 'Cassetino Stripe', 'Leopard', 'Dolce Vita', 'Mayfair']

GROUPS = {
    'metallic': {'designs': {'Imperial', 'Damask', 'Tropicana'},
                 'construction': 'Machine Made', 'materials': '100% Polyester',
                 'features': ['Smooth Pile', 'Metallic Yarn']},
    'washable': {'designs': {'Cotswold Fantasia', 'Hedgerovia', 'Sarabande', 'Dandy',
                             'Cassetino Stripe', 'Leopard'},
                 'construction': 'Digitally Printed', 'materials': '100% Polyester',
                 'features': ['Washable', 'Anti Slip', 'Flat Weave']},
    'dolce':    {'designs': {'Dolce Vita'},
                 'construction': 'Machine Made', 'materials': '100% Polypropylene',
                 'features': ['Durable']},
    'shaggy':   {'designs': {'Mayfair'},
                 'construction': 'Table Tufted', 'materials': 'Acrylic & Polyester',
                 'features': ['Shaggy', 'Textured Pile', 'Soft']},
}

INTROS = {
    'Imperial': 'Imperial is the flagship of the metallic collection, a regal ornamental design woven with metallic effect yarns in nine rich colourways.',
    'Damask': 'Damask reworks the grandest of classical patterns in metallic effect polyester yarns, glamour with a plush, smooth pile.',
    'Tropicana': 'Tropicana lets lush, exotic foliage loose across a shimmering metallic ground, glamour with a botanical twist.',
    'Cotswold Fantasia': 'Cotswold Fantasia reimagines the English country garden through a flamboyant lens, a swirling, romantic floral in painterly colour.',
    'Hedgerovia': 'Hedgerovia winds the British hedgerow into a lush, storybook botanical, fully washable and made for real homes.',
    'Sarabande': 'Sarabande moves to a stately rhythm, an opulent scrolling pattern with the grandeur of a courtly dance.',
    'Dandy': 'Dandy pairs rich, moody grounds with gilded ornamental detail, classic House Llewelyn-Bowen swagger for the floor.',
    'Cassetino Stripe': 'Cassetino Stripe is a theatrical take on the classic stripe, bold bands of colour laid down with a flamboyant flourish.',
    'Leopard': 'Leopard is animal print with maximalist attitude, a prowling pattern in bold, dramatic colour.',
    'Dolce Vita': 'Dolce Vita brings crisp, graphic pattern in easy monochrome palettes, machine made in hard wearing polypropylene.',
    'Mayfair': 'Mayfair is pure indulgence underfoot, a table tufted shaggy with an irresistibly soft, deep textured pile.',
}

WASH_LEADS = {
    'Cassetino Stripe': 'is a theatrical take on the classic stripe, bold bands of colour laid down with a flamboyant flourish',
    'Cotswold Fantasia': 'reimagines the English country garden through a flamboyant lens, a swirling, romantic floral in painterly colour',
    'Dandy': 'pairs a rich, moody ground with gilded ornamental detail, classic House Llewelyn-Bowen swagger for the floor',
    'Hedgerovia': 'winds the British hedgerow into a lush, storybook botanical brimming with leaves and blooms',
    'Leopard': 'is animal print with maximalist attitude, a prowling leopard pattern in bold, dramatic colour',
    'Sarabande': 'moves to a stately rhythm, an opulent scrolling pattern with the grandeur of a courtly dance',
}
WASH_TAIL = ('Digitally printed on a soft flat weave pile, it is fully machine washable and '
             'finished with an anti slip backing, so it is as practical as it is striking. '
             'Ideal for living rooms, bedrooms, hallways and busy family spaces.')


def clean_dashes(s):
    return re.sub(r'\s*[\u2013\u2014]\s*', ', ', s)


def norm_size(label):
    parts = [p.strip() for p in str(label).split('x')]
    return ' x '.join(str(int(p)) for p in parts)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = next((wb[n] for n in ('Product Data', 'Sheet1') if n in wb.sheetnames), wb.active)
    hdr = [c.value for c in ws[1]]
    idx = {h.split('\n')[0].strip(): i for i, h in enumerate(hdr) if h}

    def g(r, k):
        v = r[idx[k]]
        return v.strip() if isinstance(v, str) else v

    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[idx['Product Code']]]

    group_of = {}
    for gk, gv in GROUPS.items():
        for d in gv['designs']:
            group_of[d] = gk

    # group variant rows by colourway code, keeping sheet order
    colourways, order = {}, []
    for r in rows:
        code = g(r, 'Product Code')
        if code not in colourways:
            colourways[code] = []
            order.append(code)
        colourways[code].append(r)

    by_design = {}
    for code in order:
        rs = colourways[code]
        design = str(g(rs[0], 'Design')).strip()
        by_design.setdefault(design, []).append((code, rs))

    unknown = set(by_design) - set(DESIGN_ORDER)
    if unknown:
        sys.exit(f'Designs not in DESIGN_ORDER, add them with intros and a group: {sorted(unknown)}')

    designs_out = []
    for design in DESIGN_ORDER:
        if design not in by_design:
            continue
        grp = GROUPS[group_of[design]]
        entries = by_design[design]
        piles, origins, trades, rrps = set(), [], [], []
        cw_out = []
        for code, rs in entries:
            r0 = rs[0]
            # marketing colour name from the Description column
            dcol = str(g(r0, 'Description'))
            name = re.sub(r'^LLB\s*-\s*', '', dcol)
            name = re.sub(r'\s*-\s*[\d ]+x[\d ]+\s*$', '', name)
            name = re.sub(r'\s+Rug\s*$', '', name)
            colour = name[len(design):].strip() if name.startswith(design) else str(g(r0, 'Colour'))

            title = g(r0, 'Ecommerce Title')
            if not title:
                wash = ' Washable' if 'Washable' in grp['features'] else ''
                title = f'House Llewelyn-Bowen {design} {colour}{wash} Rug'

            desc = g(r0, 'Ecommerce Description')
            if desc:
                desc = clean_dashes(desc)
            elif design in WASH_LEADS:
                desc = (f'The {title} {WASH_LEADS[design]}. {WASH_TAIL}')
            elif design == 'Dolce Vita':
                c1, c2 = g(r0, 'Colour 1') or '', g(r0, 'Colour 2') or ''
                palette = (c1 + ' and ' + c2).lower() if c2 else c1.lower()
                desc = (f'The {title} brings a little la dolce vita home, crisp graphic pattern '
                        f'in an easy {palette} palette. Machine made in durable 100% polypropylene '
                        f'with a short, easy care pile, it stands up to the busiest rooms while '
                        f'keeping its graphic good looks. Ideal for living rooms, dining rooms and hallways.')
            else:
                desc = ''

            sizes = []
            for r in rs:
                label = norm_size(g(r, 'Size (cm)'))
                piles.add(g(r, 'Pile Height (cm)'))
                o = {'Turkey': 'Turkiye'}.get(g(r, 'Country of Origin'), g(r, 'Country of Origin'))
                if o not in origins:
                    origins.append(o)
                trades.append(float(g(r, 'Wholesale Price (GBP)')))
                rrps.append(float(g(r, 'RRP (GBP)')))
                sizes.append({
                    'size': label,
                    'shape': g(r, 'Shape'),
                    'runner': label == '61 x 230',
                    'variant': g(r, 'Variant Code'),
                    'rrp': round(float(g(r, 'RRP (GBP)')), 2),
                    'trade': round(float(g(r, 'Wholesale Price (GBP)')), 2),
                    'rrpEur': round(float(g(r, 'RRP (EURO)')), 2),
                    'tradeEur': round(float(g(r, 'Wholesale Price (EUR)')), 2),
                })
            cw_out.append({'code': code, 'colour': colour, 'title': title, 'desc': desc,
                           'c1': g(r0, 'Colour 1') or '', 'c2': g(r0, 'Colour 2') or '', 'sizes': sizes})

        pile_vals = sorted(float(p) for p in piles)
        if len(pile_vals) == 1:
            pile = int(pile_vals[0]) if pile_vals[0] == int(pile_vals[0]) else pile_vals[0]
        else:
            pile = f'{pile_vals[0]:g} to {pile_vals[-1]:g}'
        designs_out.append({
            'name': design,
            'intro': INTROS[design],
            'construction': grp['construction'],
            'materials': grp['materials'],
            'pile': pile,
            'origin': ' and '.join(origins),
            'features': grp['features'],
            'fromRRP': min(rrps),
            'fromTrade': min(trades),
            'colourways': cw_out,
        })

    path = f'{ROOT}/data/brand_data.json'
    data = json.load(open(path))
    data['LLB'] = {'designs': designs_out}
    json.dump(data, open(path, 'w'), indent=1, ensure_ascii=False)

    n_cw = sum(len(d['colourways']) for d in designs_out)
    n_v = sum(len(c['sizes']) for d in designs_out for c in d['colourways'])
    print(f'LLB: {len(designs_out)} designs, {n_cw} colourways, {n_v} variants -> merged into data/brand_data.json')


if __name__ == '__main__':
    main()
